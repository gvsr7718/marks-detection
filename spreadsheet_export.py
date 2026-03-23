"""
Spreadsheet Export Module
Generates styled Excel files with combined descriptive + MCQ marks.

Export format:
S.No | HT No | Q1a | Q1b | Q2a | Q2b | Q3a | Q3b | Q4a | Q4b | Q5a | Q5b | Q6a | Q6b | Descriptive Total | MCQ Score | Total
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import List, Dict


# Styles
HEADER_FILL = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# Column headers matching the user's required format
HEADERS = [
    "S.No", "HT No / Roll No",
    "Q1a", "Q1b", "Q2a", "Q2b", "Q3a", "Q3b",
    "Q4a", "Q4b", "Q5a", "Q5b", "Q6a", "Q6b",
    "Descriptive Total", "MCQ Score", "Total"
]


def _apply_headers(ws):
    """Apply styled headers to the worksheet."""
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = max(12, len(str(cell.value)) + 4)


def _apply_row_style(ws, row_idx, is_alt):
    """Apply alternating row styles."""
    for cell in ws[row_idx]:
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN
        if is_alt:
            cell.fill = ALT_FILL


def generate_combined_excel(entries: List[Dict]) -> bytes:
    """
    Generate combined Excel report with descriptive marks + MCQ score.
    
    Each entry dict should have:
    {
        "ht_no": "23241A6302",
        "marks": {
            "q1a": 3, "q1b": 4, "q2a": 2, "q2b": 2,
            "q3a": 5, "q3b": 0, "q4a": 0, "q4b": 0,
            "q5a": 0, "q5b": 0, "q6a": 0, "q6b": 0
        },
        "descriptive_total": 17,
        "mcq_score": 7,
        "total": 24
    }
    
    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Marks Report"
    
    _apply_headers(ws)
    
    for i, entry in enumerate(entries, 1):
        marks = entry.get("marks", {})
        
        descriptive_total = entry.get("descriptive_total", 0)
        mcq_score = entry.get("mcq_score", 0)
        total = entry.get("total", descriptive_total + mcq_score)
        
        row = [
            i,
            entry.get("ht_no", ""),
            marks.get("q1a", 0), marks.get("q1b", 0),
            marks.get("q2a", 0), marks.get("q2b", 0),
            marks.get("q3a", 0), marks.get("q3b", 0),
            marks.get("q4a", 0), marks.get("q4b", 0),
            marks.get("q5a", 0), marks.get("q5b", 0),
            marks.get("q6a", 0), marks.get("q6b", 0),
            descriptive_total,
            mcq_score,
            total
        ]
        
        ws.append(row)
        
        row_idx = i + 1
        _apply_row_style(ws, row_idx, is_alt=(i % 2 == 0))
        
        # Bold the total columns
        ws.cell(row=row_idx, column=15).font = TOTAL_FONT  # Descriptive Total
        ws.cell(row=row_idx, column=16).font = TOTAL_FONT  # MCQ Score
        ws.cell(row=row_idx, column=17).font = TOTAL_FONT  # Grand Total
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
